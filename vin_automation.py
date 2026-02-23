"""
vin_automation.py - Main orchestrator for the VIN Automation System.

Reads a VIN from Pinnacle Professional, decodes it via VINMatchPro,
pulls parts pricing from Car-Part.com, and exports results to CSV/Excel.
"""

import argparse
import csv
import os
import re
import sys
import time
from datetime import datetime

import openpyxl
import requests

from pinnacle_reader import JABReader, read_vin_from_pinnacle, validate_vin
from vinmatchpro_decoder import decode as decode_vin
from carpart_scraper import (
    load_config as carpart_load_config,
    search as search_parts,
    search_single_part,
    _get_homepage_hidden_fields,
)
from mvr_reader import open_mvr_and_read_parts, get_open_mvr_titles
from ebay_scraper import create_driver as ebay_create_driver, close_driver as ebay_close_driver, search_sold as ebay_search_sold


# Map Pinnacle/Hollander keywords → Car-Part.com search terms
_CARPART_NAME_MAP = [
    # Doors
    (r'\bfront door\b',             'Front Door (see also Door Shell, Front)'),
    (r'\brear door\b',              'Rear Door (side-see also Door Shell, Rear)'),
    (r'\bback door\b',              'Back Door(above bmpr-see also Door Shell, Back)'),
    # Lighting
    (r'\bheadlamp\b',               'Headlight Assembly'),
    (r'\bhead lamp\b',              'Headlight Assembly'),
    (r'\btail lamp\b',              'Tail Light'),
    (r'\btaillamp\b',               'Tail Light'),
    (r'\bpark lamp\b',              'Parking/Turn Signal Light, Front'),
    (r'\bturn signal\b',            'Parking/Turn Signal Light, Front'),
    # Drivetrain
    (r'\btrans(?:mission)?\s*at\b', 'Transmission'),
    (r'\btrans(?:mission)?\s*mt\b', 'Transmission'),
    (r'\btransmission\b',           'Transmission'),
    (r'\btransfer case motor\b',    'Transfer Case Electric Motor'),
    (r'\btransfer case\b',          'Transfer Case'),
    (r'\bturbo.supercharger\b',     'Turbocharger/Supercharger'),
    (r'\bturbocharger\b',           'Turbocharger/Supercharger'),
    (r'\bturbo\b',                  'Turbocharger/Supercharger'),
    # Engine / electronics
    (r'\bengine\b',                 'Engine'),
    (r'\bchassis brain box\b',      'Chassis Control Computer (not Engine)'),
    (r'\bengine brain box\b',       'Engine Computer'),
    # Body
    (r'\btrans\.?\s*crossmember\b', 'Transmission Crossmember'),
    (r'\bfender\b',                 'Fender'),
    (r'\bhoo[dt]\b',                'Hood'),
    (r'\bradiator\b',               'Radiator'),
]


def _clean_search_term(term: str) -> str:
    """Normalize a Pinnacle/Hollander category name to a Car-Part.com search term.

    Steps:
    1. Strip everything after a semicolon or comma-detail suffix
    2. Remove leading L/R/Left/Right side prefix  ("L Front Door…" → "Front Door…")
    3. Remove parenthetical descriptions          ("Front Door (electric…)" → "Front Door")
    4. Remove trailing specs/codes
    5. Map known keywords to Car-Part.com names
    """
    # Strip after semicolon
    term = term.split(';')[0]
    # Strip ", ID XXXX…" suffixes
    term = re.sub(r',\s*ID\s+\S+.*$', '', term)
    # Strip trailing Pinnacle part-type codes (3+ digits)
    term = re.sub(r'\s+\d{3,}\w*\s*$', '', term)
    # Remove leading side indicators: "L ", "R ", "Left ", "Right "
    term = re.sub(r'^\s*(?:L|R|Left|Right)\s+', '', term, flags=re.IGNORECASE)
    # Remove trailing side indicators: ", L.", ", R.", "halogen, L.", etc.
    term = re.sub(r',?\s+(?:L|R|LH|RH)\.?\s*$', '', term, flags=re.IGNORECASE)
    # Remove parenthetical detail after the part name
    term = re.sub(r'\s*\(.*', '', term)
    # Strip remaining trailing specs after a comma
    term = re.sub(r',.*$', '', term)
    term = term.strip().rstrip(',').strip()

    # Map to Car-Part.com name
    term_lower = term.lower()
    for pattern, carpart_name in _CARPART_NAME_MAP:
        if re.search(pattern, term_lower):
            return carpart_name

    return term


def load_config():
    """Load output settings from config.ini."""
    import configparser
    cfg = configparser.ConfigParser()
    cfg.read('config.ini')
    return {
        'output_dir': cfg.get('output', 'directory', fallback='output'),
        'format': cfg.get('output', 'format', fallback='csv').lower(),
    }


def export_csv(rows: list[dict], filepath: str):
    """Write results to a CSV file."""
    if not rows:
        print('No data to export.')
        return

    fieldnames = list(rows[0].keys())
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    print(f'CSV saved to {filepath}')


def export_excel(rows: list[dict], filepath: str):
    """Write results to an Excel (.xlsx) file."""
    if not rows:
        print('No data to export.')
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Parts Results'

    fieldnames = list(rows[0].keys())
    ws.append(fieldnames)
    for row in rows:
        ws.append([row.get(k, '') for k in fieldnames])

    wb.save(filepath)
    print(f'Excel file saved to {filepath}')


def export_excel_unpriced(rows: list[dict], filepath: str, vehicle: dict):
    """Write the unpriced-parts pricing report to a formatted Excel file."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if not rows:
        print('No data to export.')
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Pricing Report'

    HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
    HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
    ALT_FILL    = PatternFill('solid', fgColor='D6E4F0')
    MONEY_FMT   = '"$"#,##0.00'
    CENTER      = Alignment(horizontal='center', vertical='center')
    THIN        = Border(
        bottom=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
    )

    # ── Vehicle info header block ──────────────────────────────────────────
    year  = vehicle.get('year', '')
    make  = vehicle.get('make', '')
    model = vehicle.get('model', '')
    trim  = vehicle.get('trim', '')
    vin   = rows[0].get('vin', '')

    ws.merge_cells('A1:M1')
    title_cell = ws['A1']
    title_cell.value = f'Pricing Report — {year} {make} {model} {trim}'.strip()
    title_cell.font  = Font(bold=True, color='FFFFFF', size=13)
    title_cell.fill  = PatternFill('solid', fgColor='1F4E79')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24

    ws.merge_cells('A2:M2')
    vin_cell = ws['A2']
    vin_cell.value     = f'VIN: {vin}'
    vin_cell.font      = Font(italic=True, color='404040', size=10)
    vin_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 16

    # ── Column headers (row 3) ─────────────────────────────────────────────
    EBAY_FILL = PatternFill('solid', fgColor='385723')   # dark green for eBay cols

    headers = [
        ('Part Name',            28,  HEADER_FILL),
        ('Stock #',              14,  HEADER_FILL),
        ('Grade',                 8,  HEADER_FILL),
        ('Location',             14,  HEADER_FILL),
        ('Hollander #',          14,  HEADER_FILL),
        # Car-Part.com columns
        ('CarPart Avg',          13,  HEADER_FILL),
        ('CarPart Low',          13,  HEADER_FILL),
        ('CarPart #',            11,  HEADER_FILL),
        # eBay columns
        ('eBay Avg (OEM Sold)',  18,  EBAY_FILL),
        ('eBay Low (OEM Sold)',  18,  EBAY_FILL),
        ('eBay #',               10,  EBAY_FILL),
        # Notes
        ('Notes',                30,  HEADER_FILL),
    ]

    for col, (label, width, hfill) in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=label)
        cell.font      = HEADER_FONT
        cell.fill      = hfill
        cell.alignment = CENTER
        cell.border    = THIN
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[3].height = 18

    # ── Data rows ──────────────────────────────────────────────────────────
    for i, row in enumerate(rows):
        r = i + 4
        fill = ALT_FILL if i % 2 == 1 else None

        def _cell(col, value, fmt=None, align=None):
            c = ws.cell(row=r, column=col, value=value)
            if fill:
                c.fill = fill
            if fmt:
                c.number_format = fmt
            if align:
                c.alignment = align
            c.border = THIN
            return c

        def _parse_price(val):
            if not val:
                return None
            try:
                return float(str(val).replace('$', '').replace(',', ''))
            except ValueError:
                return None

        _cell(1, row.get('part_name', ''))
        _cell(2, row.get('stock_num', ''), align=CENTER)
        _cell(3, row.get('grade', ''),     align=CENTER)
        _cell(4, row.get('location', ''))
        _cell(5, row.get('hollander', ''), align=CENTER)

        # Car-Part.com
        avg = _parse_price(row.get('avg_price'))
        low = _parse_price(row.get('low_price'))
        _cell(6,  avg, fmt=MONEY_FMT, align=CENTER)
        _cell(7,  low, fmt=MONEY_FMT, align=CENTER)
        _cell(8,  row.get('listing_count', 0), align=CENTER)

        # eBay
        ebay_avg = _parse_price(row.get('ebay_avg_price'))
        ebay_low = _parse_price(row.get('ebay_low_price'))
        _cell(9,  ebay_avg, fmt=MONEY_FMT, align=CENTER)
        _cell(10, ebay_low, fmt=MONEY_FMT, align=CENTER)
        _cell(11, row.get('ebay_listing_count', 0), align=CENTER)

        _cell(12, row.get('notes', ''))

        ws.row_dimensions[r].height = 15

    # ── Freeze panes below headers ─────────────────────────────────────────
    ws.freeze_panes = 'A4'

    wb.save(filepath)
    print(f'Excel report saved to {filepath}')


def process_vin(vin: str) -> list[dict]:
    """Decode a VIN and fetch parts, returning combined result rows."""
    print(f'\n{"="*60}')
    print(f'Processing VIN: {vin}')
    print(f'{"="*60}')

    # Step 1: Decode VIN
    print('Decoding VIN via VINMatchPro...')
    try:
        vehicle = decode_vin(vin)
        print(f"  Vehicle: {vehicle.get('year', '?')} {vehicle.get('make', '?')} {vehicle.get('model', '?')}")
    except Exception as exc:
        print(f'  ERROR decoding VIN: {exc}')
        return []

    # Step 2: Search for parts
    print('Searching Car-Part.com for parts...')
    try:
        parts = search_parts(vehicle)
        print(f'  Found {len(parts)} part listing(s).')
    except Exception as exc:
        print(f'  ERROR searching parts: {exc}')
        parts = []

    # Combine vehicle info with each part row
    rows = []
    for part in parts:
        row = {
            'vin': vin,
            'year': vehicle.get('year', ''),
            'make': vehicle.get('make', ''),
            'model': vehicle.get('model', ''),
            'trim': vehicle.get('trim', ''),
            'engine': vehicle.get('engine', ''),
        }
        row.update(part)
        rows.append(row)

    # If no parts found, still output the vehicle info
    if not rows:
        rows.append({
            'vin': vin,
            'year': vehicle.get('year', ''),
            'make': vehicle.get('make', ''),
            'model': vehicle.get('model', ''),
            'trim': vehicle.get('trim', ''),
            'engine': vehicle.get('engine', ''),
            'part_name': '',
            'price': '',
            'vendor': '',
            'location': '',
            'grade': '',
        })

    return rows


def main():
    parser = argparse.ArgumentParser(
        description='VIN Automation System for Fitz Auto Parts'
    )
    parser.add_argument(
        '--vin',
        type=str,
        help='Process a specific VIN instead of reading from Pinnacle',
    )
    parser.add_argument(
        '--vins',
        type=str,
        nargs='+',
        help='Process multiple VINs',
    )
    parser.add_argument(
        '--format',
        choices=['csv', 'excel'],
        help='Override output format (csv or excel)',
    )
    parser.add_argument(
        '--unpriced',
        action='store_true',
        help=(
            'Open the selected vehicle\'s MVR, read un-priced parts with Hollander '
            'interchange numbers, look up market prices on Car-Part.com, and export '
            'a suggested-price report.'
        ),
    )
    args = parser.parse_args()

    config = load_config()
    output_dir = config['output_dir']
    output_format = args.format or config['format']

    os.makedirs(output_dir, exist_ok=True)

    # --unpriced workflow: open MVR, read Parts tab, price each un-priced part
    if args.unpriced:
        print('Reading selected VIN from Pinnacle...')
        try:
            reader = JABReader()
            vin = reader.read_selected_vin()
            print(f'  VIN: {vin}')
        except RuntimeError as exc:
            print(f'ERROR: {exc}', file=sys.stderr)
            sys.exit(1)

        print('Decoding VIN...')
        try:
            vehicle = decode_vin(vin)
            print(f"  Vehicle: {vehicle.get('year', '?')} {vehicle.get('make', '?')} {vehicle.get('model', '?')}")
        except Exception as exc:
            print(f'ERROR decoding VIN: {exc}', file=sys.stderr)
            sys.exit(1)

        print('Opening MVR (double-clicking selected row)...')
        pre_click_titles = get_open_mvr_titles()
        try:
            reader.open_selected_vehicle()
        except RuntimeError as exc:
            print(f'ERROR: {exc}', file=sys.stderr)
            sys.exit(1)

        print('Waiting for MVR to load...')
        time.sleep(3)

        print('Reading un-priced parts from MVR Parts tab...')
        try:
            parts = open_mvr_and_read_parts(pre_click_titles=pre_click_titles)
            print(f'  Found {len(parts)} un-priced part(s) with Hollander numbers.')
        except RuntimeError as exc:
            print(f'ERROR reading MVR parts: {exc}', file=sys.stderr)
            sys.exit(1)

        if not parts:
            print('No un-priced parts with Hollander numbers found. Nothing to export.')
            sys.exit(0)

        print('Looking up market prices on Car-Part.com...')
        carpart_cfg = carpart_load_config()
        zip_code = carpart_cfg['zip_code']

        session = requests.Session()
        from carpart_scraper import HEADERS
        session.headers.update(HEADERS)
        hidden = _get_homepage_hidden_fields(session)

        print('Starting eBay browser...')
        ebay_driver = ebay_create_driver()

        rows = []
        try:
            for part in parts:
                raw_term = part.get('category', '') or part['part_name']
                search_term = _clean_search_term(raw_term) if raw_term else ''
                print(f"  [{parts.index(part)+1}/{len(parts)}] {search_term or part['description'][:60]}")

                # Car-Part.com lookup
                if search_term:
                    result = search_single_part(
                        search_term,
                        vehicle.get('year', ''),
                        vehicle.get('make', ''),
                        vehicle.get('model', ''),
                        zip_code,
                        session=session,
                        hidden_fields=hidden,
                    )
                else:
                    result = {'avg_price': None, 'low_price': None, 'listing_count': 0}

                cp_avg_str = f"${result['avg_price']:.2f}" if result['avg_price'] is not None else ('$Call only' if result['listing_count'] else '')
                cp_low_str = f"${result['low_price']:.2f}" if result['low_price'] is not None else ''
                print(f"    CarPart: {cp_avg_str or 'No listings'}  ({result['listing_count']} listings)")

                # eBay OEM sold listings lookup
                ebay_result = {'avg_price': None, 'low_price': None, 'listing_count': 0}
                if search_term:
                    try:
                        ebay_result = ebay_search_sold(
                            search_term,
                            vehicle.get('year', ''),
                            vehicle.get('make', ''),
                            vehicle.get('model', ''),
                            ebay_driver,
                            oem_only=True,
                        )
                    except Exception as exc:
                        print(f'    eBay warning: {exc}')

                eb_avg_str = f"${ebay_result['avg_price']:.2f}" if ebay_result['avg_price'] is not None else ''
                print(f"    eBay OEM: {eb_avg_str or 'No listings'}  ({ebay_result['listing_count']} listings)")

                if not search_term:
                    notes = 'Description unclear – manual review'
                elif result['listing_count'] == 0 and ebay_result['listing_count'] == 0:
                    notes = 'No listings found'
                else:
                    notes = ''

                rows.append({
                    'vin':                vin,
                    'year':               vehicle.get('year', ''),
                    'make':               vehicle.get('make', ''),
                    'model':              vehicle.get('model', ''),
                    'stock_num':          part['stock_num'],
                    'hollander':          part['hollander'],
                    'part_name':          search_term or part['description'][:60],
                    'grade':              part['grade'],
                    'location':           part['location'],
                    'avg_price':          cp_avg_str,
                    'low_price':          cp_low_str,
                    'listing_count':      result['listing_count'],
                    'ebay_avg_price':     eb_avg_str,
                    'ebay_low_price':     f"${ebay_result['low_price']:.2f}" if ebay_result['low_price'] is not None else '',
                    'ebay_listing_count': ebay_result['listing_count'],
                    'notes':              notes,
                })
        finally:
            ebay_close_driver(ebay_driver)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filepath = os.path.join(output_dir, f'unpriced_{timestamp}.xlsx')
        export_excel_unpriced(rows, filepath, vehicle)
        print(f'\nDone. {len(rows)} part(s) written to {filepath}')

        import subprocess
        subprocess.Popen(['start', '', filepath], shell=True)
        return

    # Determine which VINs to process
    vins = []

    if args.vins:
        vins = [v.upper().strip() for v in args.vins]
    elif args.vin:
        vins = [args.vin.upper().strip()]
    else:
        # Read from Pinnacle Professional
        print('Reading VIN from Pinnacle Professional...')
        try:
            vin = read_vin_from_pinnacle()
            print(f'  Found VIN: {vin}')
            vins = [vin]
        except RuntimeError as exc:
            print(f'ERROR: {exc}', file=sys.stderr)
            sys.exit(1)

    # Validate all VINs
    for vin in vins:
        if not validate_vin(vin):
            print(f'WARNING: "{vin}" does not look like a valid VIN. Skipping.')
            vins.remove(vin)

    if not vins:
        print('No valid VINs to process.')
        sys.exit(1)

    # Process each VIN
    all_rows = []
    for vin in vins:
        rows = process_vin(vin)
        all_rows.extend(rows)

    # Export results
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    if output_format == 'excel':
        filepath = os.path.join(output_dir, f'parts_{timestamp}.xlsx')
        export_excel(all_rows, filepath)
    else:
        filepath = os.path.join(output_dir, f'parts_{timestamp}.csv')
        export_csv(all_rows, filepath)

    print(f'\nDone. Processed {len(vins)} VIN(s), {len(all_rows)} total row(s).')


if __name__ == '__main__':
    main()
